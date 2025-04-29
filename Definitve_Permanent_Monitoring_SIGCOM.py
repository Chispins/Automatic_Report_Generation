


from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer
import os
import time
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl.utils.exceptions import SheetTitleException
import re
#  import shutil, os, psutil
def autostart():
    import shutil, os, psutil

    # Get current user
    user = os.getlogin()

    # Path to startup folder
    startup_path = r"C:\Users\Usuario\AppData\Roaming\Microsoft\Windows\Start Menu\Programs\Startup"

    # Path to the executable (in dist folder created by PyInstaller)
    exe_path = "dist/Definitve_Permanent_Monitoring_SIGCOM.exe"

    # Target path in startup folder
    target_path = os.path.join(startup_path, "SIGCOM_Monitor.exe")

    try:
        # Copy instead of move
        shutil.copy2(exe_path, target_path)
        print(f"Successfully added to startup: {target_path}")
    except Exception as error:
        print(f"Error adding to startup: {error}")


def configurar_hoja_activa(root_dir):
    for foldername, _, filenames in os.walk(root_dir):
        if "DEVENGADO.xlsx" in filenames:
            nombre_carpeta = os.path.basename(foldername).upper()
            archivo_excel = os.path.join(foldername, "DEVENGADO.xlsx")

            try:
                wb = openpyxl.load_workbook(archivo_excel, keep_vba=True)

                if nombre_carpeta in wb.sheetnames:
                    # Configurar la hoja deseada como activa
                    wb.active = wb[nombre_carpeta]
                    wb.save(archivo_excel)
                    print(f"Configurada hoja activa: {archivo_excel} -> '{nombre_carpeta}'")
                else:
                    print(f"Advertencia: No existe la hoja '{nombre_carpeta}' en {archivo_excel}")

                wb.close()

            except SheetTitleException:
                print(f"Error: Nombre de hoja inválido en {archivo_excel}")
            except Exception as e:
                print(f"Error procesando {archivo_excel}: {str(e)}")
directorio_raiz = "RUTA/A/TUS/CARPETAS"
configurar_hoja_activa(directorio_raiz)

base_path = r"\\10.5.130.24\Abastecimiento\Compartido Abastecimiento\Otros\SIGCOM"
work_directory = r'C:\Users\Usuario\Downloads'  # Where data files are located
meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']


class FileMonitorHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        print(f'File created: {event.src_path}')
        self.process_file(event.src_path)

    def on_moved(self, event):
        if event.is_directory:
            return
        print(f'File renamed: {event.src_path} -> {event.dest_path}')
        self.process_file(event.dest_path)

    def on_modified(self, event):
        if event.is_directory:
            # Check if this is a directory we're monitoring
            verificar_carpetas(event.src_path)
            return
        print(f'File modified: {event.src_path}')
        self.process_file(event.src_path)

    def check_required_files_exist(self):
        """Check if Codigos_Clasificador_Compilado.xlsx exists in the shared network folder."""
        # New path for required file
        required_file_path = r"\\10.5.130.24\Abastecimiento\Compartido Abastecimiento\Otros\SIGCOM\No Borrar\Codigos_Clasificador_Compilado.xlsx"

        if not os.path.exists(required_file_path):
            print(f"ERROR: Codigos_Clasificador_Compilado.xlsx is missing in the shared network folder")
            # Try to diagnose the problem
            try:
                # Check if the parent directory exists
                parent_dir = os.path.dirname(required_file_path)
                if not os.path.exists(parent_dir):
                    print(f"Parent directory does not exist: {parent_dir}")
                else:
                    print(f"Parent directory exists. Checking contents...")
                    files = os.listdir(parent_dir)
                    print(f"Files in directory: {files}")

                # Check if we can access the network share at all
                base_share = r"\\10.5.130.24\Abastecimiento"
                if not os.path.exists(base_share):
                    print(f"Cannot access network share: {base_share}")
                else:
                    print(f"Network share is accessible.")

            except Exception as e:
                print(f"Error diagnosing file access: {str(e)}")

            return False

        print(f"Required file found: {required_file_path}")
        # Check if we can actually read the file
        try:
            with open(required_file_path, 'rb') as f:
                f.read(1)  # Try to read 1 byte
            print("File is readable")
        except Exception as e:
            print(f"File exists but cannot be read: {str(e)}")
            return False

        return True

    def check_devengado_exists(self, folder_path):
        """Check if DEVENGADO file exists in the same folder as the BASE DISTRIBUCION file."""
        if not os.path.exists(folder_path):
            return False

        devengado_files = [f for f in os.listdir(folder_path) if f.startswith("DEVENGADO") and f.endswith(".xlsx")]
        if not devengado_files:
            print(f"ERROR: No DEVENGADO file found in {folder_path}")
            return False

        print(f"Found DEVENGADO file: {devengado_files[0]} in {folder_path}")
        return True




    def check_modified_exists(self, file_path):
        """Check if a modified version of the file already exists."""
        folder = os.path.dirname(file_path)
        filename = os.path.basename(file_path)
        modified_file = os.path.join(folder, f"Modified_{filename}")

        if os.path.exists(modified_file):
            print(f"WARNING: Modified file already exists: {modified_file}")
            print("Processing skipped to avoid overwriting existing modified file.")
            return True
        return False

    def process_file(self, file_path):
        # Only process Excel files starting with "BASE DISTRIBUCION GASTO GENERAL"
        filename = os.path.basename(file_path)
        folder_path = os.path.dirname(file_path)

        # Skip temporary, modified, or exported files
        if filename.startswith('~$') or filename.startswith('Modified_') or filename.startswith(
                'Exported_') or not filename.endswith('.xlsx'):
            return

        # Check if modified file already exists
        if self.check_modified_exists(file_path):
            return

        # Check if required files exist before processing
        if not self.check_required_files_exist():
            print("Processing aborted due to missing Codigos_Clasificador_Compilado.xlsx.")
            return

        # Check if DEVENGADO file exists in the same folder
        if not self.check_devengado_exists(folder_path):
            print("Processing aborted due to missing DEVENGADO file in the same folder.")
            return

        try:
            print(f"Processing Excel file: {file_path}")
            update_excel_with_xlwings(file_path)
            print(f"Processing complete. Modified file saved.")
        except Exception as e:
            print(f"Error processing file: {str(e)}")

            def check_modified_exists(self, file_path):
                """Check if a modified version of the file already exists."""
                folder = os.path.dirname(file_path)
                filename = os.path.basename(file_path)
                modified_file = os.path.join(folder, f"Modified_{filename}")

                if os.path.exists(modified_file):
                    print(f"WARNING: Modified file already exists: {modified_file}")
                    print("Processing skipped to avoid overwriting existing modified file.")
                    return True
                return False

def verificar_carpetas(carpeta_modificada=None):
    """Verifies year/month folders and processes Excel files."""
    for año in range(2024, 2041):
        for mes in meses:
            carpeta_mes = os.path.join(base_path, str(año), mes)

            # Skip if folder doesn't exist
            if not os.path.exists(carpeta_mes):
                continue

            # Skip if a specific folder is being checked and doesn't match
            if carpeta_modificada and carpeta_modificada != carpeta_mes:
                continue

            # List files in the folder
            archivos = os.listdir(carpeta_mes)

            # Find matching files
            base_files = [a for a in archivos if a.endswith('.xlsx') and a.startswith('BASE DISTRIBUCION GASTO GENERAL')]
            devengado_files = [a for a in archivos if a.startswith("DEVENGADO") and a.endswith(".xlsx")]

            # Skip if no matching files
            if not devengado_files or not base_files:
                continue

            for archivo in base_files:
                file_path = os.path.join(carpeta_mes, archivo)

                # Create an instance of the handler to use its methods
                handler = FileMonitorHandler()

                # Check if modified file already exists
                if handler.check_modified_exists(file_path):
                    continue

                # Verify required files exist
                if not handler.check_required_files_exist():
                    print("Codigos_Clasificador_Compilado.xlsx missing, cannot process.")
                    continue

                try:
                    print(f"Processing Excel file: {file_path}")
                    update_excel_with_xlwings(file_path)
                    print(f"Processing complete. Modified file saved.")
                except Exception as e:
                    print(f"Error processing file: {str(e)}")


def format_sigfe_code(code):
    if not isinstance(code, str):
        return code
    parts = code.split('.')
    if len(parts) < 3:
        return code
    result = f"{parts[0]}.{parts[1]}"
    for i in range(2, len(parts)):
        section = parts[i].strip()
        if len(section) >= 2:
            result += f".{section[-2:]}"
        else:
            result += f".{section}"
    return result


def trim_at_first_repeat(text):
    if not isinstance(text, str):
        return text
    words = text.split()
    seen = set()
    result = []
    for word in words:
        if word in seen:
            break
        seen.add(word)
        result.append(word)
    return ' '.join(result)


def calculate_subasignaciones_amounts(exported):
    exported["MONTO_SUBASIGNACIONES"] = exported["MONTO "].copy()
    base_codes = exported.index[exported["Subasignaciones"] == 1].tolist()
    code_to_base = {}
    for code in exported.index:
        matching_bases = [base for base in base_codes if str(code).startswith(str(base))]
        if matching_bases:
            code_to_base[code] = max(matching_bases, key=len)
        else:
            code_to_base[code] = None
    base_sums = {}
    for base in base_codes:
        related_codes = [code for code, mapped_base in code_to_base.items() if mapped_base == base]
        base_sums[base] = exported.loc[related_codes, "MONTO "].sum()
    for code in exported.index:
        base = code_to_base.get(code)
        if base is not None:
            exported.loc[code, "MONTO_SUBASIGNACIONES"] = base_sums[base]
    exported["MONTO_SUBASIGNACIONES"] = (exported["MONTO_SUBASIGNACIONES"] - exported["MONTO "]) * (
        exported["Subasignaciones"])
    return exported


def process_data(folder_path):
    """Process data using files from work_directory and the specified folder."""
    original_dir = os.getcwd()
    try:
        # Find DEVENGADO file in the folder path
        devengado_files = [f for f in os.listdir(folder_path) if f.startswith("DEVENGADO") and f.endswith(".xlsx")]
        if not devengado_files:
            raise FileNotFoundError(f"No DEVENGADO file found in {folder_path}")
        devengado_file = os.path.join(folder_path, devengado_files[0])

        shared_path = r"\\10.5.130.24\Abastecimiento\Compartido Abastecimiento\Otros\SIGCOM\No Borrar"
        categorias_codigos_file = os.path.join(shared_path, "Codigos_Clasificador_Compilado.xlsx")

        # Process categorias_codigos data from work_directory
        # os.chdir(work_directory)
        # categorias_codigos = pd.read_excel("Codigos_Clasificador_Compilado.xlsx")
        categorias_codigos = pd.read_excel(categorias_codigos_file)
        categorias_codigos["Subasignaciones"] = categorias_codigos["Cod_SIGFE"].str.contains("y").astype(int)
        categorias_codigos.loc[categorias_codigos["Subasignaciones"] == 1, "Cod_SIGFE"] = categorias_codigos[
            categorias_codigos["Subasignaciones"] == 1]["Cod_SIGFE"].apply(
            lambda x: x.split()[0] if isinstance(x, str) and ' ' in x else x.replace("y", "") if isinstance(x,
                                                                                                            str) else x)

        # Format SIGFE codes
        categorias_codigos["Cod_SIGFE"] = categorias_codigos["Cod_SIGFE"].apply(format_sigfe_code)
        categorias_codigos["Cod_SIGFE"] = categorias_codigos["Cod_SIGFE"].str.replace(".", "")
        categorias_codigos["Cod_SIGFE"] = categorias_codigos["Cod_SIGFE"].str[2:]


        # This part is just to read the sheetname of the month in the folder name
        folder_name = os.path.basename(os.getcwd())
        devengado_excel = pd.ExcelFile(devengado_file)
        nombre_hojas = devengado_excel.sheet_names

        sheet_name = None
        for hoja in nombre_hojas:
            if re.search(folder_name, hoja, re.IGNORECASE):
                sheet_name = hoja
                break

        # If no matching sheet found, use default behavior
        if sheet_name:
            print(f"Using matching sheet: '{sheet_name}' for folder: '{folder_name}'")
            devengado = pd.read_excel(devengado_file, skiprows=5, header=0, sheet_name=sheet_name)
        else:
            print(f"No matching sheet found for folder '{folder_name}'. Using default sheet.")
            devengado = pd.read_excel(devengado_file, skiprows=5, header=0)





        devengado = pd.read_excel(devengado_file, skiprows=5, header=0)

        devengado = devengado.dropna(subset=["NOMBRE PROVEEDOR"])

        devengado["item_conv"] = devengado.iloc[:, 9].copy()
        devengado["Cod_SIGFE"] = devengado.iloc[:, 9].copy()
        devengado["item_conv"] = devengado["item_conv"].str.split("-").str[1].str.strip()
        devengado["Cod_SIGFE"] = devengado["Cod_SIGFE"].str.split("-").str[1].str.strip()



        # Rest of function remains the same
        merged = pd.merge(devengado, categorias_codigos, how='outer', on='Cod_SIGFE')
        # Now generate a devengado_modified_excel
        exported = merged.groupby(by="Cod_SIGFE").agg({
            "MONTO ": "sum",
            "Item_en_SIGCOM": "first",
            "Item SIGFE": "first",
            "Subasignaciones": "sum"})

        exported["Item_en_SIGCOM"] = exported["Item_en_SIGCOM"].apply(trim_at_first_repeat)
        exported["Item SIGFE"] = exported["Item SIGFE"].apply(trim_at_first_repeat)
        exported["Item_en_SIGCOM"] = exported["Item_en_SIGCOM"].str[:65]
        exported["Item SIGFE"] = exported["Item SIGFE"].str[:65]

        exported["MONTO "] = exported["MONTO "]
        exported = calculate_subasignaciones_amounts(exported)
        return exported, merged
    finally:
        # Restore original directory
        os.chdir(original_dir)


def configurar_hoja_activa(root_dir):
    for foldername, _, filenames in os.walk(root_dir):
        if "DEVENGADO.xlsx" in filenames:
            nombre_carpeta = os.path.basename(foldername).upper()
            archivo_excel = os.path.join(foldername, "DEVENGADO.xlsx")

            try:
                wb = openpyxl.load_workbook(archivo_excel, keep_vba=True)

                if nombre_carpeta in wb.sheetnames:
                    # Configurar la hoja deseada como activa
                    wb.active = wb[nombre_carpeta]
                    wb.save(archivo_excel)
                    print(f"Configurada hoja activa: {archivo_excel} -> '{nombre_carpeta}'")
                else:
                    print(f"Advertencia: No existe la hoja '{nombre_carpeta}' en {archivo_excel}")

                wb.close()

            except SheetTitleException:
                print(f"Error: Nombre de hoja inválido en {archivo_excel}")
            except Exception as e:
                print(f"Error procesando {archivo_excel}: {str(e)}")


def process_file(self, file_path):
    # First check if it's a DEVENGADO file
    if os.path.basename(file_path).upper().startswith("DEVENGADO"):
        self.configurar_hoja_activa(file_path)
        return

    # Rest of existing process_file logic...
    # [Keep all the original BASE DISTRIBUCION processing code here]


def debug_xlwings_update(file_path):
    """
    Debug function that only tests the xlwings portion of the update process
    to identify why the Modified file isn't being saved.

    Args:
        file_path: Path to the BASE DISTRIBUCION Excel file
    """
    print(f"\n=== DEBUGGING XLWINGS UPDATE ===")
    print(f"Testing file: {file_path}")

    # Check if file exists
    if not os.path.exists(file_path):
        print(f"ERROR: File does not exist: {file_path}")
        return

    # Check if the new file already exists
    folder = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    new_file = os.path.join(folder, f"Modified_{folder}")

    if os.path.exists(new_file):
        print(f"WARNING: Modified file already exists: {new_file}")
        print("Will attempt to overwrite.")

    # Create some test values (don't need real data for debugging)
    test_values = {
        "GASTO GENERAL": [(3, 21, 1000), (3, 23, 2000)],
        "SUMINISTROS": [(2, 7, 3000), (3, 20, 4000)]
    }

    app = None
    wb = None
    try:
        print("1. Starting Excel application...")
        app = xw.App(visible=False)

        print(f"2. Opening workbook: {file_path}")
        wb = xw.Book(file_path)

        print("3. Getting sheet names...")
        sheet_names = [sheet.name for sheet in wb.sheets]
        print(f"   Found sheets: {sheet_names}")

        # Try updating a few test cells in each sheet
        for sheet_name, cells in test_values.items():
            if sheet_name in sheet_names:
                print(f"4. Updating {sheet_name} sheet...")
                sheet = wb.sheets[sheet_name]
                for row, col, value in cells:
                    try:
                        print(f"   Setting cell ({row}, {col}) to {value}")
                        sheet.cells(row, col).value = value
                    except Exception as e:
                        print(f"   ERROR updating cell ({row}, {col}): {str(e)}")
            else:
                print(f"   Sheet '{sheet_name}' not found!")

        print(f"5. Saving as: {new_file}")
        wb.save(new_file)

        print("6. Checking if file was created...")
        if os.path.exists(new_file):
            print(f"   SUCCESS: File created successfully: {new_file}")
            print(f"   File size: {os.path.getsize(new_file)} bytes")
        else:
            print(f"   ERROR: File was not created: {new_file}")

    except Exception as e:
        print(f"ERROR in xlwings operations: {str(e)}")
    finally:
        print("7. Cleaning up Excel application...")
        if wb:
            try:
                wb.close()
                print("   Workbook closed")
            except Exception as e:
                print(f"   ERROR closing workbook: {str(e)}")
        if app:
            try:
                app.quit()
                print("   Excel application quit")
            except Exception as e:
                print(f"   ERROR quitting Excel: {str(e)}")

        print("=== DEBUGGING COMPLETE ===\n")


def safe_get_value(df, code, column):
    """Safely extract a value from DataFrame, returning 0 if not found."""
    try:
        return df.loc[code, column]
    except (KeyError, TypeError):
        print(f"Warning: Code '{code}' not found in data, using 0 instead.")
        return 0


def update_excel_with_xlwings(file_path):
    # Get the folder path
    folder_path = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    base_filename = os.path.splitext(filename)[0]

    # Process data with the folder path
    exported, merged = process_data(folder_path)

    # Save the exported DataFrame to Excel
    exported_file = os.path.join(folder_path, f"Exported_{base_filename}.xlsx")
    try:
        print(f"Saving exported data to: {exported_file}")
        exported.to_excel(exported_file)
        print(f"Exported data saved as {exported_file}")
    except Exception as e:
        print(f"Error saving exported data: {str(e)}")

    # Save merged devengado
    merged_file = os.path.join(folder_path, f"Modified_DEVENGADO_{base_filename}.xlsx")
    try:
        print(f"Saving merged data to: {merged_file}")
        merged.to_excel(merged_file)
        print(f"Merged data saved as {merged_file}")
    except Exception as e:
        print(f"Error saving merged data: {str(e)}")

    # Extract values for GASTO GENERAL sheet using safe_get_value
    gas = safe_get_value(exported, "0503", "MONTO ")
    mantencion_jardines = safe_get_value(exported, "0803", "MONTO ")
    eq_computo = safe_get_value(exported, "0607", "MONTO ")
    agua = safe_get_value(exported, "0502", "MONTO ")
    servicio_energia = safe_get_value(exported, "0501", "MONTO ")

    servicios_informaticos = safe_get_value(exported, "0906", "MONTO ")
    mantenimiento_maquinaria = safe_get_value(exported, "060501", "MONTO ")
    mantenimiento_correctivo = safe_get_value(exported, "060604", "MONTO ")
    mantenimiento_preventivo = safe_get_value(exported, "060603", "MONTO ")
    servicio_laboratorio = safe_get_value(exported, "12990201", "MONTO ")
    seguridad_y_vigilancia = safe_get_value(exported, "0802", "MONTO ")
    servicio_de_aseo = safe_get_value(exported, "0801", "MONTO ")
    gasto_medicamentos_x_servicio = safe_get_value(exported, "129919", "MONTO ")
    servicio_lavanderia = safe_get_value(exported, "040701", "MONTO ")
    otros_gastos = safe_get_value(exported, "12990202", "MONTO ")
    mantenimiento_y_reparacion_vehiculo = safe_get_value(exported, "0602", "MONTO_SUBASIGNACIONES")
    mantenimiento_planta_fisica = safe_get_value(exported, "0601", "MONTO ")
    pasaje_traslado_pacientes = safe_get_value(exported, "129910", "MONTO ")
    mantenimiento_mueble_y_enserses = safe_get_value(exported, "0603", "MONTO ")
    cursos_capacitacion = safe_get_value(exported, "1102", "MONTO ") + safe_get_value(exported, "1103", "MONTO_SUBASIGNACIONES")
    coloc_adulto_mayor = safe_get_value(exported, "12990603", "MONTO ")
    material_curacion = safe_get_value(exported, "040403", "MONTO ")

    # Extract values for SUMINISTROS sheet using safe_get_value
    combustible_vehiculos = safe_get_value(exported, "0301", "MONTO ")
    material_medico_quirurjico = safe_get_value(exported, "0405", "MONTO_SUBASIGNACIONES")
    material_oficina = safe_get_value(exported, "0401", "MONTO ")
    materiales_informaticos = safe_get_value(exported, "0409", "MONTO ")
    gasto_total_medicamentos = safe_get_value(exported, "040401", "MONTO_SUBASIGNACIONES")
    productos_quimicos = safe_get_value(exported, "040302", "MONTO ")
    gasas_medicinales = safe_get_value(exported, "040301", "MONTO ")
    material_mantencion = safe_get_value(exported, "0410", "MONTO ")
    material_aseo = safe_get_value(exported, "040702", "MONTO ")
    material_casino = safe_get_value(exported, "0408", "MONTO ")
    vestuario_y_calzado = safe_get_value(exported, "0202", "MONTO_SUBASIGNACIONES")
    alimentacion_viveres_pacientes = safe_get_value(exported, "010101", "MONTO ")
    alimentacion_viveres_funcionarios = safe_get_value(exported, "010102", "MONTO ")
    combustible_y_lubricante_calefaccion = safe_get_value(exported, "0303", "MONTO ")
    material_curacion = safe_get_value(exported, "040403", "MONTO ")

    app = None
    wb = None
    try:
        # Open the BASE DISTRIBUCION file from the No Borrar folder
        print("Opening BASE DISTRIBUCION file from No Borrar folder...")
        base_distribucion_path = r"\\10.5.130.24\Abastecimiento\Compartido Abastecimiento\Otros\SIGCOM\No Borrar\BASE DISTRIBUCION GASTO GENERAL Y SUMINISTROS.xlsx"

        if not os.path.exists(base_distribucion_path):
            raise FileNotFoundError(f"File not found: {base_distribucion_path}")

        app = xw.App(visible=False)
        wx = xw.Book(base_distribucion_path)

        # Initialize a new workbook for copying sheets
        wb = xw.Book()

        for hoja in wx.sheets:
            hoja.api.Copy(Before=wb.sheets[0].api)

        # Delete the default sheet if it exists
        if "Hoja1" in [sheet.name for sheet in wb.sheets]:
            wb.sheets["Hoja1"].delete()

        sheet_names = [sheet.name for sheet in wb.sheets]

        # Update GASTO GENERAL sheet
        if "GASTO GENERAL" in sheet_names:
            sheet = wb.sheets["GASTO GENERAL"]
            print("Updating GASTO GENERAL sheet...")
            sheet.cells(3, 21).value = gas
            sheet.cells(3, 23).value = mantencion_jardines
            sheet.cells(6, 25).value = eq_computo
            sheet.cells(3, 54).value = agua
            sheet.cells(3, 58).value = servicio_energia

            sheet.cells(3, 5).value = servicios_informaticos
            sheet.cells(3, 27).value = mantenimiento_maquinaria
            sheet.cells(3, 37).value = mantenimiento_correctivo
            sheet.cells(3, 41).value = mantenimiento_preventivo
            sheet.cells(3, 64).value = servicio_laboratorio
            sheet.cells(3, 73).value = seguridad_y_vigilancia
            sheet.cells(3, 56).value = servicio_de_aseo
            sheet.cells(3, 60).value = gasto_medicamentos_x_servicio
            sheet.cells(2, 66).value = servicio_lavanderia
            sheet.cells(6, 16).value = otros_gastos
            sheet.cells(6, 35).value = mantenimiento_y_reparacion_vehiculo
            sheet.cells(89, 34).value = mantenimiento_planta_fisica
            sheet.cells(62, 15).value = coloc_adulto_mayor
            sheet.cells(3, 48).value = pasaje_traslado_pacientes
            sheet.cells(3, 31).value = mantenimiento_mueble_y_enserses
            sheet.cells(3, 18).value = cursos_capacitacion
            sheet.cells(3, 24).value = material_curacion
        else:
            print("Warning: 'GASTO GENERAL' sheet not found!")

        # Update SUMINISTROS sheet
        if "SUMINISTROS" in sheet_names:
            sheet = wb.sheets["SUMINISTROS"]
            print("Updating SUMINISTROS sheet...")
            sheet.cells(2, 7).value = combustible_vehiculos
            sheet.cells(3, 20).value = material_medico_quirurjico
            sheet.cells(3, 28).value = material_oficina
            sheet.cells(3, 31).value = materiales_informaticos
            sheet.cells(3, 39).value = gasto_total_medicamentos
            sheet.cells(3, 46).value = productos_quimicos
            sheet.cells(2, 11).value = gasas_medicinales
            sheet.cells(2, 33).value = material_mantencion
            sheet.cells(2, 37).value = material_aseo
            sheet.cells(2, 43).value = material_casino
            sheet.cells(2, 50).value = vestuario_y_calzado
            sheet.cells(2, 55).value = alimentacion_viveres_pacientes
            sheet.cells(2, 58).value = alimentacion_viveres_funcionarios
            sheet.cells(2, 5).value = combustible_y_lubricante_calefaccion
            sheet.cells(3, 24).value = material_curacion
        else:
            print("Warning: 'SUMINISTROS' sheet not found!")

        # Save the modified file
        new_file = os.path.join(folder_path, f"Modified_{filename}")
        wb.save(new_file)
        print(f"Excel file updated and saved as {new_file}")
    finally:
        # Ensure Excel closes properly
        if wb:
            try:
                wb.close()
            except:
                pass
        if app:
            try:
                app.quit()
            except:
                pass


def iniciar_monitoreo():
    """Setup monitoring for the base directory and all year/month subfolders."""
    print(f"Monitoring active in: {base_path} (Ctrl+C to stop)")
    event_handler = FileMonitorHandler()
    observer = Observer()
    observer.schedule(event_handler, base_path, recursive=True)
    observer.start()

    try:
        # Process all existing folders when starting
        verificar_carpetas()

        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    #autostart()
    iniciar_monitoreo()


# Ojo con subasignaciones, tengo duda de que funcionen bien